# Working with spreadsheets using openpyxl module. Must have example.xlsx in working directory.
import openpyxl

# Excel/excel type spreadsheet workbook
wb = openpyxl.load_workbook('example.xlsx')

# /////////////////////// Sheets

# Get all sheet names
all_sheets = wb.get_sheet_names()

# Get certain sheet by given name value
sheet = wb.get_sheet_by_name('Sheet1')
# print(sheet.title)  # Access sheet's attributes like title

active = wb.get_active_sheet()

# ///////////////////////// Cells

cA1 = sheet['A1']  # or sheet.cell(row=1, column=1)
# print(cA1.value)  # Access cell's values, row, column, coordinate etc

# Access each row from column in steps of 2
for x in range(1, 8, 2):
    print(x, sheet.cell(row=x, column=2).value)

max_rows = sheet.max_row
max_cols = sheet.max_column

print(max_rows)

